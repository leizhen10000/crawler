#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
# @Time    : 2018/3/3 10:18
# @Author  : Lei Zhen
# @Contract: leizhen8080@gmail.com
# @File    : extact_contact.py
# @Software: PyCharm
# code is far away from bugs with the god animal protecting
    I love animals. They taste delicious.
             ┏┓   ┏┓
            ┏┛┻━━━┛┻┓
            ┃  ☃    ┃
            ┃ ┳┛  ┗┳ ┃
            ┃      ┻  ┃
            ┗━┓     ┏━┛
              ┃     ┗━━━┓
              ┃ 神兽保佑 ┣┓
              ┃ 永无BUG┏┛
              ┗┓┓┏━┳┓┏┛
               ┃┫┫ ┃┫┫
               ┗┻┛ ┗┻┛
"""
# 首先连接 mongodb
import json

import pymongo
from openpyxl import Workbook

connector = 'mongodb://{username}:{password}@{host}/{db}'.format(
    username='jr_qb', password='7hIrimaUSuLP9LPTXZRy', host='10.139.154.117', db='jr_qb')
connection = pymongo.MongoClient(connector)
collection_name = 'wd_api_mobilephone_getdatav2_response'

# 连接数据库
db = connection.get_database('jr_qb')
print('连接数据库 ' + json.dumps(db.name, indent=4, sort_keys=True) + ' 成功')

# 连接表对象
collection = db.get_collection(collection_name)
print('连接集合 ' + json.dumps(collection.name, indent=4, sort_keys=True) + ' 成功')

# 获取数据
post = collection.find_one({'uid': 200885519})
data_list = post['data']['data_list']
# 获取 call_time
item_list = []  # 所有的 call_item
for data in data_list:
    tel_data_list = data['teldata']
    for tel_data in tel_data_list:
        call_item = tel_data['items']
        item_list.extend(call_item)
call_time_list = []
for item in item_list:
    call_time_list.append(item['call_time'])
# 获取 send_time
msg_list = []  # 所有 send_item
for data in data_list:
    msg_data_list = data['msgdata']
    for msg_data in msg_data_list:
        msg_item = msg_data['items']
        msg_list.extend(msg_item)
send_time_list = []
for msg in msg_list:
    send_time_list.append(msg['send_time'])
# 断开数据连接
connection.close()

# 把 list 数据存入 excel
# 写入
wb1 = Workbook()

ws1 = wb1.get_active_sheet()
print('默认的excel工作簿名称', ws1.title)

# 设置单元格的值
# ws1['d3'].value = 4
# ws1.cell(row=3, column=1).value = 5 #两种方式写入
ws1['a1'].value = 'call_time'
ws1['b1'].value = 'send_time'
for row in range(2, len(call_time_list) + 2):
    ws1.cell(row=row, column=1).value = call_time_list[row - 2]
for row in range(2, len(send_time_list) + 2):
    ws1.cell(row=row, column=2).value = send_time_list[row - 2]

wb1.save(filename='./new_file.xlsx')
