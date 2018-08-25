#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
# @Time    : 2018/5/17 14:34
# @Author  : Lei Zhen
# @Contract: leizhen8080@gmail.com
# @File    : read_data.py
# @Software: PyCharm
# code is far away from bugs with the god animal protecting
    I love animals. They taste delicious.
             ┏┓   ┏┓
            ┏┛┻━━━┛┻┓
            ┃  ☃    ┃
            ┃ ┳┛  ┗┳ ┃
            ┃      ┻  ┃
            ┗━┓     ┏━┛
              ┃     ┗━━━┓
              ┃ 神兽保佑 ┣┓
              ┃ 永无BUG┏┛
              ┗┓┓┏━┳┓┏┛
               ┃┫┫ ┃┫┫
               ┗┻┛ ┗┻┛
"""
import pymysql
from openpyxl import Workbook


def read_from_mysql():
    connect = pymysql.connect(
        host='localhost', user='root', password='root', charset='utf8',
        cursorclass=pymysql.cursors.DictCursor

    )
    try:
        with connect.cursor() as cursor:
            cursor.execute("select * from mybatis.college;")
            data = cursor.fetchall()
            return data
    except Exception as e:
        raise Exception('查询数据报错，请重新检查')
        raise e
    finally:
        connect.close()


def write2_excel(college_info_list):
    """写入excel"""

    wb1 = Workbook()

    ws1 = wb1.get_sheet_by_name('Sheet')
    print('默认的excel工作簿名称', ws1.title)

    # 设置单元格的值
    ws1['a1'].value = '院校名称'
    ws1.cell(row=2, column=1).value = '学历层次'
    ws1.cell(row=3, column=1).value = '院校属性'

    for college_index in range(0, len(college_info_list)):
        row = college_index + 1
        ws1.cell(row=row, column=1).value = college_info_list[college_index]['name']
        ws1.cell(row=row, column=2).value = college_info_list[college_index]['level']
        ws1.cell(row=row, column=3).value = college_info_list[college_index]['feature']

    wb1.save(filename='../source/excel/college.xlsx')

    wb1.close()


if __name__ == '__main__':
    data = read_from_mysql()
    write2_excel(data)
