#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
# @Time    : 2018/1/5 19:39
# @Author  : Lei Zhen
# @Contract: leizhen8080@gmail.com
# @File    : excel_demo.py
# @Software: PyCharm
# code is far away from bugs with the god animal protecting
    I love animals. They taste delicious.
             ┏┓   ┏┓
            ┏┛┻━━━┛┻┓
            ┃  ☃    ┃
            ┃ ┳┛  ┗┳ ┃
            ┃      ┻  ┃
            ┗━┓     ┏━┛
              ┃     ┗━━━┓
              ┃ 神兽保佑 ┣┓
              ┃ 永无BUG┏┛
              ┗┓┓┏━┳┓┏┛
               ┃┫┫ ┃┫┫
               ┗┻┛ ┗┻┛
"""
# openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


def deal_with_openpyxl():
    wb = load_workbook(filename='../source/excel/test.xlsx')

    sheets = wb.get_sheet_names()
    print('所有工作表的名称：', sheets)
    ws = wb.get_sheet_by_name('工作表1')
    sheet1 = sheets[0]

    # 获取表格所有行和列，两者都是可迭代的
    rows = ws.rows
    cols = ws.columns

    # 行迭代
    content = []
    for row in rows:
        line = [col.value for col in row]
        content.append(line)

    # 通过坐标读取值
    print(ws['B2'].value)
    print(ws.cell(row=2, column=2).value)

    wb.close()

    # 写入
    wb1 = Workbook()

    ws1 = wb1.get_active_sheet()
    print('默认的excel工作簿名称', ws1.title)

    # 设置单元格的值
    ws1['d3'].value = 4
    ws1.cell(row=3, column=1).value = 5

    ws2 = wb1.create_sheet(title='new_sheet')
    for row in range(1, 100):
        for col in range(1, 10):
            ws2.cell(row=row, column=col).value = col

    ws3 = wb1.create_sheet(title='Data')
    for row in range(10, 20):
        for rol in range(27, 54):
            _ = ws3.cell(column=col, row=row,
                         value="{0}".format(get_column_letter(col)))

    col_c = ws2['c']
    # print('c 列所有数据: ', col_c)
    col_range = ws2['c:d']
    # print('c，d 两列数据: ', col_range)
    row_10 = ws2[10]
    # print('第 10 行所有数据: ', row_10)
    row_range = ws2[3:4]
    # print('从第三行到第四行所有数据： ', row_range)

    wb1.save(filename='../source/excel/new_file.xlsx')


if __name__ == '__main__':
    deal_with_openpyxl()
